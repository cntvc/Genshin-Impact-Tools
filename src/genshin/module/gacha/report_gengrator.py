"""
gacha data export
"""
import abc
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from genshin.config import settings
from genshin.core import logger
from genshin.core.function import save_json
from genshin.module.gacha.data_transform import convert_to_uigf
from genshin.module.gacha.metadata import (GACHA_QUERY_TYPE_IDS, GACHA_QUERY_TYPE_NAMES,
                                           GACHA_TYPE_DICT)

_alignment = Alignment(horizontal="left", vertical="center")
_title_font = Font(name="微软雅黑", bold=True, color="8e8e8e")
_content_font = Font(name="微软雅黑", color="8e8e8e")
_star_5_font = Font(name="微软雅黑", color="bd6932")
_star_4_font = Font(name="微软雅黑", color="a256e1")
_star_3_font = Font(name="微软雅黑", color="c4c2bf")


class AbstractGenerator(metaclass=abc.ABCMeta):
    def __init__(self, data: Optional[dict], uid: Optional[str]) -> None:
        self.data = data
        self.uid = uid

    @abc.abstractmethod
    def generator(self):
        pass

    @abc.abstractmethod
    def status(self):
        pass


class XLSXGenerator(AbstractGenerator):
    def __init__(self, data: Optional[dict], uid: Optional[str]) -> None:
        super().__init__(data, uid)

    def status(self):
        return True

    def generator(self):
        logger.debug("开始生成XLSX报告")
        workbook_path = Path(settings.USER_DATA_PATH, self.uid, "抽卡数据总览.xlsx").as_posix()
        workbook = Workbook()

        statistical_sheet = workbook.active
        statistical_sheet.title = "数据总览"
        statistical_sheet.column_dimensions["A"].width = 16
        statistical_sheet.column_dimensions["E"].width = 16
        statistical_sheet.column_dimensions["I"].width = 16
        statistical_sheet.column_dimensions["J"].width = 16
        statistical_sheet.column_dimensions["K"].width = 16
        statistical_sheet.column_dimensions["L"].width = 16

        start_col = 8
        # 列标题-卡池名
        for index_head_col, name in enumerate(["项目"] + GACHA_QUERY_TYPE_NAMES, start=1):
            cell = statistical_sheet.cell(row=1, column=index_head_col, value=name)
            cell.alignment = _alignment
            cell.font = _title_font
        # 列标题-5星详情
        for index_head_col, name in enumerate(GACHA_QUERY_TYPE_NAMES, start=1):
            cell = statistical_sheet.cell(row=1, column=start_col + index_head_col, value=name)
            cell.alignment = _alignment
            cell.font = _title_font

        col_name = ["抽卡总数", "5星出货次数", "出5星平均次数", "保底内抽数"]
        # 行标题-统计项目
        for row, name in enumerate(col_name, 2):
            cell = statistical_sheet.cell(row=row, column=1, value=name)
            cell.font = _title_font

        for statistical_col, gacha_type_id in enumerate(GACHA_QUERY_TYPE_IDS, start=2):
            gacha_type_list = self.data["list"][gacha_type_id][:]
            gacha_type_name = GACHA_TYPE_DICT[gacha_type_id]
            logger.debug("开始写入 {}, 共 {} 条数据", gacha_type_name, len(gacha_type_list))
            excel_header = ["抽卡总数", "时间", "名称", "类别", "星级", "祈愿类型", "保底内抽数"]
            worksheet = workbook.create_sheet(gacha_type_name)
            worksheet.column_dimensions["B"].width = 22
            worksheet.column_dimensions["C"].width = 16
            worksheet.column_dimensions["F"].width = 16
            worksheet.column_dimensions["G"].width = 16
            for index_head_col, name in enumerate(excel_header, 1):
                cell = worksheet.cell(row=1, column=index_head_col, value=name)
                cell.alignment = _alignment
                cell.font = _title_font

            total_counter = 0
            pity_counter = 0
            star_5_list = []
            for index_row, gacha in enumerate(gacha_type_list, start=2):
                time = gacha["time"]
                name = gacha["name"]
                item_type = gacha["item_type"]
                rank_type = int(gacha["rank_type"])
                gacha_type = gacha["gacha_type"]
                gacha_type_name = GACHA_TYPE_DICT.get(gacha_type, "")
                total_counter = total_counter + 1
                pity_counter = pity_counter + 1

                excel_data = [
                    total_counter,
                    time,
                    name,
                    item_type,
                    rank_type,
                    gacha_type_name,
                    pity_counter,
                ]
                # worksheet.append(excel_data)

                for index_col, data in enumerate(excel_data, start=1):
                    cell = worksheet.cell(row=index_row, column=index_col, value=data)
                    cell.alignment = _alignment
                    if rank_type == 5:
                        cell.font = _star_5_font
                    elif rank_type == 4:
                        cell.font = _star_4_font
                    elif rank_type == 3:
                        cell.font = _star_3_font
                if rank_type == 5:
                    star_5_list.append("{}@{}抽".format(name, pity_counter))
                    pity_counter = 0

            five_average = "-"
            if len(star_5_list):
                five_average = (total_counter - pity_counter) / len(star_5_list)
                five_average = round(five_average, 2)

            # 5 星数据统计
            statistical_data = [total_counter, len(star_5_list), five_average, pity_counter]
            for statistical_row, data in enumerate(statistical_data, start=2):
                cell = statistical_sheet.cell(
                    row=statistical_row, column=statistical_col, value=data
                )
                cell.alignment = _alignment
                cell.font = _content_font

            # 5星抽卡详情
            for star_5_row, five_start in enumerate(star_5_list, start=2):
                cell = statistical_sheet.cell(
                    row=star_5_row, column=start_col + statistical_col - 1, value=five_start
                )
                cell.alignment = _alignment
                cell.font = _star_5_font

        workbook.save(workbook_path)
        workbook.close()
        logger.debug("XLSX文件写入完成")
        return True


class UIGFGenerator(AbstractGenerator):
    def __init__(self, data: Optional[dict], uid: Optional[str]) -> None:
        super().__init__(data, uid)
        self._uigf = {}

    def generator(self):
        self._uigf = convert_to_uigf(self.data)
        path = Path(settings.USER_DATA_PATH, self.uid, "gacha_data_uigf.json")
        if not save_json(path, self._uigf):
            logger.error("保存UIGF格式失败")
        self._uigf_to_xlsx()

    def status(self):
        return settings.FLAG_GENERATOR_UIGF

    def _uigf_to_xlsx(self):
        workbook_path = Path(settings.USER_DATA_PATH, self.uid, "抽卡数据总览.xlsx").as_posix()
        workbook = load_workbook(workbook_path)
        worksheet = workbook.create_sheet("原始数据")
        header = [
            "count",
            "gacha_type",
            "id",
            "item_id",
            "item_type",
            "lang",
            "name",
            "rank_type",
            "time",
            "uid",
            "uigf_gacha_type",
        ]
        for head_col, item in enumerate(header, start=1):
            cell = worksheet.cell(row=1, column=head_col, value=item)
            cell.alignment = _alignment
            cell.font = _title_font

        for row, gacha_item in enumerate(self._uigf["list"], start=1):
            count = gacha_item.get("count", "")
            gacha_type = gacha_item.get("gacha_type", "")
            id = gacha_item.get("id", "")
            item_id = gacha_item.get("item_id", "")
            item_type = gacha_item.get("item_type", "")
            lang = gacha_item.get("lang", "")
            name = gacha_item.get("name", "")
            rank_type = gacha_item.get("rank_type", "")
            time_str = gacha_item.get("time", "")
            uid = gacha_item.get("uid", "")
            uigf_gacha_type = gacha_item.get("uigf_gacha_type", "")
            excel_data = [
                count,
                gacha_type,
                id,
                item_id,
                item_type,
                lang,
                name,
                rank_type,
                time_str,
                uid,
                uigf_gacha_type,
            ]
            for col, data in enumerate(excel_data, start=1):
                cell = worksheet.cell(row=row, column=col, value=data)
                cell.alignment = _alignment
                cell.font = _content_font
        workbook.save(workbook_path)
        workbook.close()


class ReportManager:
    def __init__(self, data: Optional[dict], uid: Optional[str]) -> None:
        self.data = data
        self.uid = uid
        self.generators: List[AbstractGenerator] = []

    def generator_report(self):
        logger.info("开始生成抽卡报告")
        for generator in self.generators:
            if not generator.status():
                continue
            generator.data = self.data
            generator.uid = self.uid
            generator.generator()
        logger.info("生成抽卡报告任务完成")

    def add_generator(self, generator: AbstractGenerator):
        self.generators.append(generator)


xlsx_generator = XLSXGenerator(None, None)
uigf_generator = UIGFGenerator(None, None)
report = ReportManager(None, None)
report.add_generator(xlsx_generator)
report.add_generator(uigf_generator)
